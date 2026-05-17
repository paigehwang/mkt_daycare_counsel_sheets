"""
Parse 4 본부 xlsx files into structured JSON for the dashboard.
"""
import openpyxl
import re
import json
import os
from collections import defaultdict
from datetime import datetime, timedelta

RAW = '/home/claude/dashboard/raw'
OUT = '/home/claude/dashboard/data.json'

CENTER_ALIASES = {
    '김해 봉황점': '김해점',
    '대구서구점': '대구 서구점',
    '양산 물금점': '양산점',
    '천안 서북구점': '천안점',
    '광주 봄날점 사회복지사': '광주 봄날점',
    '봄날점': '광주 봄날점',
    '광주 북구점 사회복지사': '광주 북구점',
    # Map HQ/team identifiers to their bonbu group
    '수도권1': '수도권1본부 (기타)',
    '성장기획팀, 성장지원팀': '본사 성장팀',
}

# No exclusions — include 배원/김미가 etc. as their own categories
EXCLUDE_CENTERS = set()

def parse_center(raw):
    if not raw:
        return None
    raw = str(raw)
    m = re.search(r'\[([^\]]+)\]', raw)
    if not m:
        return None
    inside = m.group(1).strip()
    inside = re.sub(r'\s*(센터장|복지팀장|본부)\s*$', '', inside)
    inside = re.sub(r'\([^)]*\)', '', inside).strip()
    if not inside:
        return None
    inside = CENTER_ALIASES.get(inside, inside)
    if inside in EXCLUDE_CENTERS:
        return None
    return inside

def parse_month(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m')
    # Excel serial number (float/int)
    if isinstance(val, (int, float)):
        try:
            d = datetime(1899, 12, 30) + timedelta(days=float(val))
            return d.strftime('%Y-%m')
        except (ValueError, OverflowError):
            return None
    s = str(val).strip()
    for pat in [r'(\d{4})-(\d{1,2})', r'(\d{4})\.(\d{1,2})', r'(\d{4})/(\d{1,2})']:
        m = re.match(pat, s)
        if m:
            return f"{m.group(1)}-{int(m.group(2)):02d}"
    return None

def normalize_channel(val):
    if val is None:
        return '기타'
    s = str(val).strip().replace(' ', '')
    return s if s else '기타'

rows = []
sheet_stats = {}
detail_rows = []  # full per-row data for detail view

for fname in sorted(os.listdir(RAW)):
    if not fname.endswith('.xlsx'):
        continue
    bonbu = fname.replace('.xlsx', '')
    wb = openpyxl.load_workbook(os.path.join(RAW, fname), data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
        col = {h: i for i, h in enumerate(headers)}
        ai, bi, ci = col.get('담당자'), col.get('상담일'), col.get('유입경로')
        phone_i = col.get('연락처')
        if ai is None or bi is None or ci is None:
            continue
        count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= max(ai, bi, ci):
                continue
            center = parse_center(row[ai])
            month = parse_month(row[bi])
            channel = normalize_channel(row[ci])
            if not center or not month:
                continue
            # date as YYYY-MM-DD if possible
            raw_date = row[bi]
            if isinstance(raw_date, datetime):
                date_str = raw_date.strftime('%Y-%m-%d')
            elif isinstance(raw_date, (int, float)):
                try:
                    d = datetime(1899, 12, 30) + timedelta(days=float(raw_date))
                    date_str = d.strftime('%Y-%m-%d')
                except (ValueError, OverflowError):
                    date_str = month + '-01'
            else:
                ds = str(raw_date).strip()
                m = re.match(r'(\d{4})[-./](\d{1,2})[-./](\d{1,2})', ds)
                date_str = f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}" if m else month + '-01'
            phone = ''
            if phone_i is not None and phone_i < len(row) and row[phone_i]:
                phone = str(row[phone_i]).strip()
            rows.append({'bonbu': bonbu, 'sheet': sheet_name, 'center': center, 'month': month, 'channel': channel})
            detail_rows.append({
                'b': bonbu, 's': sheet_name, 'c': center,
                'd': date_str, 'm': month, 'ch': channel, 'p': phone
            })
            count += 1
        sheet_stats[f"{bonbu}/{sheet_name}"] = count

print("=== Parse stats ===")
for k, v in sheet_stats.items():
    print(f"  {k}: {v}")
print(f"Total: {len(rows)}")

counts = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int))))
channels = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(int)))))

for r in rows:
    counts[r['bonbu']][r['center']][r['month']][r['sheet']] += 1
    channels[r['bonbu']][r['center']][r['month']][r['sheet']][r['channel']] += 1

def conv(d):
    if isinstance(d, (defaultdict, dict)):
        return {k: conv(v) for k, v in d.items()}
    return d

months = sorted({r['month'] for r in rows})
centers_by_bonbu = defaultdict(set)
for r in rows:
    centers_by_bonbu[r['bonbu']].add(r['center'])
centers_by_bonbu = {k: sorted(v) for k, v in centers_by_bonbu.items()}
all_channels = sorted({r['channel'] for r in rows})

output = {
    'counts': conv(counts),
    'channels': conv(channels),
    'months': months,
    'centers_by_bonbu': centers_by_bonbu,
    'all_channels': all_channels,
    'sheet_stats': sheet_stats,
    'totals': {'rows': len(rows)},
    'sheets': ['등급신청', '유선상담', '대면상담', '계약상담', '상담요청'],
    'detail': detail_rows,
}

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(output, f, ensure_ascii=False, separators=(',', ':'))

print(f"\nMonths: {len(months)} ({months[0]} ~ {months[-1]})")
for b, cs in centers_by_bonbu.items():
    print(f"  {b}: {len(cs)} centers - {cs}")
print(f"Channels: {all_channels}")
print(f"Output: {os.path.getsize(OUT):,} bytes")
